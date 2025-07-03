# Copyright 2025 Google LLC
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#    https://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

"""A command-line tool to manage Dialogflow CX intents.

This script interacts with Dialogflow CX agents to manage intents,
including retrieval and description generation using Vertex AI Gemini API.
"""

import argparse
import json
import logging
import re
import sys
import time
import traceback
import uuid

from google import genai
from google.api_core import client_options
from google.api_core.retry import Retry
from google.cloud import dialogflowcx_v3
from httpx import ReadTimeout
from jsonschema import ValidationError, validate
from langcodes import Language
import nltk
import openpyxl
import stopwords
from tqdm import tqdm

# Configure logging
logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s", stream=sys.stdout
)
logger = logging.getLogger(__name__)

# Default LLM
DEFAULT_LLM = "gemini-2.5-pro-exp-03-25"

# Timeout for LLM calls
LLM_TIMEOUT_SECONDS = 600

# Schemas for Gemini responses
TRAINING_PHRASE_RECOMMENDATION_SCHEMA = {
    "type": "array",
    "items": {
        "type": "object",
        "description": "Recommendation for a training phrase.",
        "properties": {
            "recommendation": {
                "type": "string",
                "enum": ["ADD", "UPDATE", "REMOVE", "RETAIN"],
                "description": "Recommendation to ADD, UPDATE, REMOVE or RETAIN the training phrase",
            },
            "explanation": {
                "type": "string",
                "description": "Explanation for the recommendation in the language of the training phrase",
            },
            "originalTrainingPhrase": {
                "type": "string",
                "description": "The original training phrase string.",
            },
            "newTrainingPhrase": {
                "type": "string",
                "description": "New or updated training phrase string.",
            },
        },
        "required": ["recommendation", "originalTrainingPhrase", "explanation"],
    },
}


class ConversationalAgentsBestPractice:
    def _get_cx_client(self, agent_name: str, client_class) -> any:
        """Returns a Dialogflow CX client for the specified API."""
        location = agent_name.split("/")[3]
        if location != "global":
            client_options_ = client_options.ClientOptions(
                api_endpoint=f"{location}-dialogflow.googleapis.com"
            )
            return client_class(client_options=client_options_)
        return client_class()

    def get_intents_client(self, agent_name: str) -> dialogflowcx_v3.IntentsClient:
        """Returns the Dialogflow CX intents client."""
        return self._get_cx_client(agent_name, dialogflowcx_v3.IntentsClient)

    def get_entity_types_client(self, agent_name: str) -> dialogflowcx_v3.EntityTypesClient:
        """Returns the Dialogflow CX entity types client."""
        return self._get_cx_client(agent_name, dialogflowcx_v3.EntityTypesClient)

    def get_intents(self, agent_name: str, language_code: str) -> list[dialogflowcx_v3.Intent]:
        """Retrieves all intents for a given Dialogflow CX agent."""
        logger.info(f"Retrieving intents for agent: {agent_name}")
        intents_client = self.get_intents_client(agent_name=agent_name)
        request = dialogflowcx_v3.ListIntentsRequest(
            parent=agent_name, language_code=language_code
        )
        intents = list(intents_client.list_intents(request=request, retry=Retry()))
        # ensure training phrases have an ID
        for intent in intents:
            for training_phrase in intent.training_phrases:
                if not training_phrase.id:
                    training_phrase.id = str(uuid.uuid4())
        logger.info(f"Retrieved {len(intents)} intents.")
        return intents

    def list_entity_types(self, agent_name: str, language_code: str) -> list[dialogflowcx_v3.EntityType]:
        """Lists all entity types for a given Dialogflow CX agent."""
        entity_types_client = self.get_entity_types_client(agent_name=agent_name)
        request = dialogflowcx_v3.ListEntityTypesRequest(
            parent=agent_name, language_code=language_code
        )
        entity_types = list(
            entity_types_client.list_entity_types(request=request, retry=Retry())
        )
        return entity_types

    def format_training_phrase_string(self, phrase: dialogflowcx_v3.Intent.TrainingPhrase) -> str:
        """Formats a single training phrase into a single string."""
        parts = []
        for part in phrase.parts:
            if part.parameter_id:
                if "sys." in part.parameter_id:
                    parts.append(f"[{part.text}](sys.{part.parameter_id})")
                else:
                    parts.append(f"[{part.text}]({part.parameter_id})")
            else:
                parts.append(part.text)
        return "".join(parts)

    def get_training_phrase_strings(self, intent: dialogflowcx_v3.Intent) -> list[str]:
        """Extracts and formats training phrases from an intent as strings."""
        return [
            self.format_training_phrase_string(phrase)
            for phrase in intent.training_phrases
        ]

    def validate_response(self, response_text: str, schema: dict[str, any]) -> list[dict[str, any]]:
        """Validates a JSON response against a schema, returns recommendations."""
        try:
            response_json = json.loads(response_text)

            if isinstance(response_json, dict) and "recommendations" in response_json:
                recommendations = response_json["recommendations"]
            elif isinstance(response_json, list):  # Assume response is directly the recommendations array
                recommendations = response_json
            else:
                logger.error(
                    f"'recommendations' key not found and response is not a list: {response_json}"
                )
                return []

            validate(instance=recommendations, schema=schema)
            return recommendations
        except json.JSONDecodeError as e:
            logger.error(f"Invalid JSON response: {e}")
            return []
        except ValidationError as e:
            logger.error(f"Schema validation error: {e}")
            return []

    def generate_with_gemini(
        self,
        prompt: str,
        model_name: str,
        generate_content_config: genai.types.GenerateContentConfig,
        response_schema: dict[str, any] | None = None,
        vertex_ai_project: str | None = None,
        vertex_ai_location: str | None = None,
        gemini_timeout: int = LLM_TIMEOUT_SECONDS,
        max_retries: int = 3,
        initial_retry_delay: int = 1,
    ) -> str:
        """Generates content with Gemini, with timeout, retries, and schema validation."""
        #http_options = genai.types.HttpOptions(timeout=gemini_timeout * 1000)
        # TODO: remove
        http_options = genai.types.HttpOptions(timeout=gemini_timeout * 1000, base_url=f"https://{vertex_ai_location}-staging-aiplatform.sandbox.googleapis.com")
        client = genai.Client(
            vertexai=True,
            project=vertex_ai_project,
            location=vertex_ai_location,
            http_options=http_options,
        )
        response_text = ""
        retry = False

        for attempt in range(max_retries + 1):
            logger.info(f"Attempt #{attempt} with model {model_name}")
            # Log token counts before generating content
            try:
                token_count_response = client.models.count_tokens(
                    model=model_name, contents=prompt
                )
            except:
                token_count_response = None
                pass
            if token_count_response:
                logger.info(
                    f"Token count for prompt: {token_count_response.total_tokens}"
                )
            else:
                logger.warning("Failed to get token count.")                             

            try:
                model_response = client.models.generate_content(
                    model=model_name,
                    contents=prompt,
                    config=generate_content_config,
                )

                response_text = model_response.text.strip(" `\n")
                response_text = re.sub(r"^[^\[]+","", response_text)                
                logger.debug(f"Model response beginning:\n{response_text[0:1000]}")
                logger.debug(f"Model response end:\n{response_text[-1000:-1]}")
                logger.debug(f"Model response length: {len(response_text)}")
                #logger.debug(f"Model response:\n{response_text}")

                if response_schema:
                    validated_response = self.validate_response(
                        response_text, response_schema
                    )
                    if validated_response:
                        return json.dumps(validated_response)
                    else:                        
                        retry = True
            
            except ReadTimeout as e:
                logger.error(
                    f"Gemini API call timed out after {gemini_timeout} seconds (Attempt {attempt+1}/{max_retries+1})."
                )
                retry = True

            except Exception as e:
                logger.error(
                    f"Gemini API Error (Attempt {attempt+1}/{max_retries+1}): {e}"
                )
                retry = True
            
            if retry:
                retry = False
                response_text = ""               
                if attempt < max_retries:
                    retry_delay = initial_retry_delay * (2**attempt)
                    logger.info(f"Retrying in {retry_delay} seconds...")
                    time.sleep(retry_delay)
                    continue
                else:
                    logger.error(
                        f"Max retries reached. Gemini API call failed after {max_retries+1} attempts."
                    )
        return response_text

    def generate_intent_description(
        self,
        intent: dialogflowcx_v3.Intent,
        language_code: str,
        model_name: str,
        vertex_ai_project: str | None = None,
        vertex_ai_location: str | None = None,
        gemini_timeout: int = LLM_TIMEOUT_SECONDS,
    ) -> str:
        """Generates an intent description using Vertex AI Gemini API."""
        logger.info(f"Generating description for intent: {intent.display_name}")
        language = Language.get(language_code).display_name("en")

        generate_content_config = genai.types.GenerateContentConfig(
            temperature=0.2, max_output_tokens=140  # Description length limit
        )
        training_phrases = self.get_training_phrase_strings(intent)
        prompt = f"""
                Generate a precise description of the users intent based on the user input in the training phrases. The description must have less than 140 characters.

                The description must be generated in {language}

                Intent Display Name:
                {intent.display_name}

                Training Phrases:
                {training_phrases}

                Description:
                """
        logger.debug(
            f"Generating description for intent {intent.name} using prompt:\n{prompt}"
        )
        return self.generate_with_gemini(
            prompt=prompt,
            model_name=model_name,
            generate_content_config=generate_content_config,
            vertex_ai_project=vertex_ai_project,
            vertex_ai_location=vertex_ai_location,
            gemini_timeout=gemini_timeout,
        )

    def get_project(self, agent_name: str) -> str:
        """Extracts the project ID from the agent name."""
        return agent_name.split("/")[1]

    def update_intent_description(
        self,
        intent: dialogflowcx_v3.Intent,
        description: str,
        agent_name: str,
        language_code: str,
    ) -> None:
        """Updates the intent description in Dialogflow CX."""
        if intent.description == description:
            logger.info(
                f"Intent {intent.display_name} already has the same description. Skipping update."
            )
            return

        intent.description = description
        update_mask = {"paths": ["description"]}
        intents_client = self.get_intents_client(agent_name=agent_name)
        request = dialogflowcx_v3.UpdateIntentRequest(
            intent=intent, language_code=language_code, update_mask=update_mask
        )
        logger.info(
            f"Updating intent: {intent.display_name} with description '{description}'"
        )
        intents_client.update_intent(request=request, retry=Retry())
        logger.info(f"Intent description updated: {intent.display_name}")

    def generate_training_phrase_recommendations(
        self,
        intent: dialogflowcx_v3.Intent,
        language_code: str,
        model_name: str,
        entity_types: list[dialogflowcx_v3.EntityType],
        vertex_ai_project: str | None = None,
        vertex_ai_location: str | None = None,
        gemini_timeout: int = LLM_TIMEOUT_SECONDS,
    ) -> list[dict[str, any]]:
        """Generates training phrase recommendations using Gemini."""
        logger.info(
            f"Generating training phrase recommendations for intent: {intent.display_name}"
        )
        project = self.get_project(agent_name=intent.name)
        intent_training_phrases = self.get_training_phrase_strings(intent)

        language = Language.get(language_code).language_name("en")

        try:
            nltk.data.find("corpora/stopwords")
        except LookupError:
            nltk.download("stopwords")

        try:
            stop_words = set(
                stopwords.safe_get_stopwords(Language.get(language_code).language)
            )
            stop_words = sorted(
                stop_words.union(set(nltk.corpus.stopwords.words(language)))
            )
        except OSError as e:
            logger.warning(
                f"Could not load stopwords for language code: {language_code}. Stopwords will not be used in prompt. Error: {e}"
            )
            stop_words = set()

        generate_content_config = genai.types.GenerateContentConfig(
            temperature=0.1,
            max_output_tokens=65535,
            response_mime_type="application/json",
            response_schema=TRAINING_PHRASE_RECOMMENDATION_SCHEMA,
        )

        entity_type_prompt_section = ""
        if intent.parameters:
            entity_type_prompt_section = "*   **Available Parameter IDs:**\n```\n"
            for parameter in intent.parameters:                
                entity_type_prompt_section += f"{parameter.id}\n"
            entity_type_prompt_section += "```"

        prompt = f"""
**Task:** Analyze the provided {language} training phrases for the intent "{intent.display_name}" and suggest improvements (ADD, UPDATE, REMOVE, RETAIN) to enhance NLU model performance.

**Recommendation Guidelines:**

*   **Recommendation Types:**
    *   **ADD**  Only add phrases if absolutely necessary. Current training phrase is empty for this suggestion. Suggest new training phrases to increase variety, coverage of user expressions. Focus on real user queries. Explain why they should be added.
    *   **UPDATE** Modify existing phrases to properly use stop words. Explain why the stop words need to be added.
    *   **REMOVE** Identify and recommend removal of redundant, ambiguous, or ineffective phrases. Explain why the phrase is not beneficial.
    *   **RETAIN** Indicate phrases that are already well-formed and contribute effectively to the intent. Provide an empty explanation.

*   **Key Considerations for Recommendations:**
    *   **Existing Training Phrases**: Training phrases with just one word, incomplete words, transcription errors, spelling or grammar mistakes must be **RETAIN**. Similar training phrases without the errors may be **ADD**.
    *   **Language:** All explanations must be in {language}. All phrases with recommendation ADD or UPDATE must be in {language}.
    *   **Clarity & Conciseness:** Phrases should be clear, concise, and directly related to the intent.
    *   **Variety & Coverage:** Ensure diverse phrasing, including different ways users ask about the intent.
    *   **Real-world Language & Stop Words:** **Prioritize natural, conversational {language}. Include common {language} stop words (articles, possessive pronouns, prepositions - see list below) where they make the phrase sound more natural and idiomatic.**
    *   **Avoid Filler Words (Unnecessary):** Exclude *unnecessary* filler words. Do not create phrases that *only* differ by filler words. However, include stop words that are essential for natural {language}.
    *   **Stop Words**: **UPDATE** all phrases which benefit from adding stop words.
    *   **Quantity (Guideline):** Aim for at least 10-20 high-quality phrases per intent.
    *   **Valid JSON:** Check that the response is valid JSON

**{language} stop words:**
```
{stop_words}
```

{entity_type_prompt_section}

**Intent Description:** {intent.description}

**Training Phrases to Analyze (with parameters indicated as [text](parameter_id)):**

{intent_training_phrases}

Training Phrase Recommendations:
"""

        logger.debug(f"Generating recommendation with prompt:\n{prompt}")

        recommendations_str = self.generate_with_gemini(
            prompt=prompt,
            model_name=model_name,
            generate_content_config=generate_content_config,
            response_schema=TRAINING_PHRASE_RECOMMENDATION_SCHEMA,
            vertex_ai_project=vertex_ai_project,
            vertex_ai_location=vertex_ai_location,
            gemini_timeout=gemini_timeout,
        )
        
        return json.loads(recommendations_str) if recommendations_str else []

    def _extract_phrase_parts(
        self, phrase_string: str, intent_parameters: list[dialogflowcx_v3.Intent.Parameter]
    ) -> list[dialogflowcx_v3.Intent.TrainingPhrase.Part]:
        """Extracts parts from a training phrase string with parameter info."""
        parts = []
        # Regex to find parameters like [text](parameterId)
        param_pattern = re.compile(r"\[([^\]]+)\]\(([^)]+)\)")
        start = 0
        for match in param_pattern.finditer(phrase_string):
            text = match.group(1)
            param_id = match.group(2)
            # Add any preceding text as a plain part
            if match.start() > start:
                parts.append(
                    dialogflowcx_v3.Intent.TrainingPhrase.Part(
                        text=phrase_string[start : match.start()]
                    )
                )
            # Add the matched parameter part
            parts.append(
                dialogflowcx_v3.Intent.TrainingPhrase.Part(
                    text=text, parameter_id=param_id
                )
            )
            start = match.end()

        # Add any remaining text as a plain part
        if start < len(phrase_string):
            parts.append(
                dialogflowcx_v3.Intent.TrainingPhrase.Part(text=phrase_string[start:])
            )
        return parts

    def _match_entities(
        self,
        text: str,
        intent_parameters: list[dialogflowcx_v3.Intent.Parameter],
        entity_types: list[dialogflowcx_v3.EntityType],
    ) -> list[dialogflowcx_v3.Intent.TrainingPhrase.Part]:
        """Automatically matches entities in a training phrase string."""
        words = text.split()
        parts = []
        i = 0
        while i < len(words):
            matched = False
            # Try matching from longest possible substring
            for j in range(len(words), i, -1):
                substring = " ".join(words[i:j])
                for parameter in intent_parameters:
                    entity_type = next(
                        (
                            et
                            for et in entity_types
                            if et.name == parameter.entity_type
                        ),
                        None,
                    )
                    if entity_type:
                        # Check for match against entity values and synonyms
                        for entity in entity_type.entities:
                            if substring.lower() == entity.value.lower() or substring.lower() in [
                                syn.lower() for syn in entity.synonyms
                            ]:
                                parts.append(
                                    dialogflowcx_v3.Intent.TrainingPhrase.Part(
                                        text=substring, parameter_id=parameter.id
                                    )
                                )
                                i = j
                                matched = True
                                break
                        if matched:
                            break
                        # Check for regex match if entity type is KIND_REGEXP
                        if entity_type.kind == dialogflowcx_v3.EntityType.Kind.KIND_REGEXP:
                            for entity in entity_type.entities:
                                try:
                                    if re.fullmatch(entity.value, substring, re.IGNORECASE):
                                        parts.append(
                                            dialogflowcx_v3.Intent.TrainingPhrase.Part(
                                                text=substring, parameter_id=parameter.id
                                            )
                                        )
                                        i = j
                                        matched = True
                                        break
                                except re.error as e:
                                    logger.debug(f"Regex error for {entity.value}: {e}")
                            if matched:
                                break
                if matched:
                    break
            if not matched:
                parts.append(
                    dialogflowcx_v3.Intent.TrainingPhrase.Part(text=words[i])
                )
                i += 1
        return parts

    def update_intent_training_phrases(
        self,
        intent: dialogflowcx_v3.Intent,
        training_phrases_strings: list[str],
        agent_name: str,
        language_code: str,
        entity_types: list[dialogflowcx_v3.EntityType],
    ) -> None:
        """Updates the intent's training phrases in Dialogflow CX."""
        training_phrase_objects = []
        for phrase_string_data in training_phrases_strings:
            # Attempt to parse parameter format first
            parts = self._extract_phrase_parts(
                phrase_string_data["text"], intent.parameters
            )
            # If no parameters were found, perform automatic matching
            #if all(not part.parameter_id for part in parts):
            #    parts = self._match_entities(
            #        phrase_string_data["text"], intent.parameters, entity_types
            #    )

            training_phrase_objects.append(
                dialogflowcx_v3.Intent.TrainingPhrase(
                    parts=parts,
                    repeat_count=phrase_string_data.get("repeatCount", 1),
                    id=phrase_string_data.get("id", None),  # Keep ID if it exists.
                )
            )

        intent.training_phrases = training_phrase_objects
        logger.info(f"Updating training phrases for intent: {intent.display_name}")
        self.update_intent(agent_name, intent, language_code)
            
        logger.info(f"Training phrases updated for intent: {intent.display_name}")

    def update_intent(self, 
                      agent_name: str, 
                      intent: dialogflowcx_v3.Intent,
                      language_code: str):
        intents_client = self.get_intents_client(agent_name=agent_name)
        update_mask = {"paths": ["training_phrases"]}
        request = dialogflowcx_v3.UpdateIntentRequest(
            intent=intent, language_code=language_code, update_mask=update_mask
        )
        try:
            intents_client.update_intent(request=request, retry=Retry())
        except Exception as e:
            logger.warning(f"Exception occured: {str(e)}")
            match = re.search(r"Parameter '([^']+)' in training phrase part is not defined in the intent.",str(e))
            if match and match.lastindex == 1:
                parameter = match.group(1)
                logger.info(f"Removing parameter {parameter} and trying again")
                for training_phrase in intent.training_phrases:
                    for part in training_phrase.parts:
                        if part.parameter_id == parameter:
                            part.parameter_id = None
                self.update_intent(agent_name, intent, language_code)
            else:
                raise e

    def create_excel_output(
        self,
        intents: list[dialogflowcx_v3.Intent],
        recommendations: dict[str, list[dict[str, any]]],
        output_file: str,
        agent_name: str,
        entity_types: list[dialogflowcx_v3.EntityType],
    ) -> None:
        """Creates and saves an Excel file with intent recommendations."""
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Intent Recommendations"

        columns = [
            {"name": "Intent Name", "hide": True},
            {"name": "Intent Display Name", "hide": True},
            {"name": "Intent Description", "width": 30, "wrap": True},
            {"name": "Training Phrase ID", "hide": True},
            {"name": "Repeat Count", "hide": True},
            {"name": "Current Training Phrase", "width": 40, "wrap": True},
            {"name": "New Training Phrase", "width": 40, "wrap": True},
            {"name": "Recommendation", "width": 16, "wrap": True},
            {"name": "Explanation", "width": 60, "wrap": True},
            {"name": "Apply Change", "width": 12},
        ]

        col_num = 1
        for column in columns:
            cell = sheet.cell(row=1, column=col_num)
            cell.value = column.get("name")
            if "width" in column:
                sheet.column_dimensions[
                    openpyxl.utils.cell.get_column_letter(col_num)
                ].width = column.get("width")
            if "hide" in column and column.get("hide"):
                sheet.column_dimensions[
                    openpyxl.utils.cell.get_column_letter(col_num)
                ].hidden = column.get("hide")
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal="center")
            col_num += 1

        row_num = 2
        for intent in intents:
            if intent.display_name == "Default Negative Intent":
                continue

            intent_name = intent.name
            intent_display_name = intent.display_name
            intent_description = intent.description
            intent_recommendations = recommendations.get(intent_name, [])
            current_training_phrases = self.get_training_phrase_strings(intent)
            # Create a map for easy lookup of original training phrases
            original_tp_map = {
                tp.id: tp_string
                for tp, tp_string in zip(
                    intent.training_phrases, current_training_phrases
                )
            }

            first_row_intent = row_num
            num_recommendations = len(intent_recommendations)

            for rec_index, rec in enumerate(intent_recommendations):
                original_phrase_id = (
                    rec.get("originalTrainingPhraseId")
                    if "originalTrainingPhraseId" in rec
                    else None
                )
                current_phrase_str = rec.get("originalTrainingPhrase", "")
                new_phrase_str = rec.get("newTrainingPhrase", "")
                recommendation = rec.get("recommendation", "")
                explanation = rec.get("explanation", "")
                repeat_count = 1
                if original_phrase_id and original_phrase_id in original_tp_map:
                    current_phrase_str = original_tp_map[original_phrase_id]
                    tp_obj = next(tp for tp in intent.training_phrases if tp.id == original_phrase_id)
                    repeat_count = tp_obj.repeat_count

                row_values = [
                    intent_name,
                    intent_display_name,
                    intent_description,
                    original_phrase_id,
                    repeat_count,
                    current_phrase_str,
                    new_phrase_str,
                    recommendation,
                    explanation,
                    "X" if recommendation != "RETAIN" else "",
                ]
                sheet.append(row_values)

                # Apply wrapping to cells where wrap is true
                for col_index, column in enumerate(columns):
                    if "wrap" in column and column["wrap"]:
                        cell = sheet.cell(row=row_num, column=col_index + 1)
                        cell.alignment = openpyxl.styles.Alignment(
                            wrap_text=True, vertical="top"
                        )
                row_num += 1

            if num_recommendations > 0:
                sheet.merge_cells(
                    start_row=first_row_intent,
                    end_row=row_num - 1,
                    start_column=1,
                    end_column=1,
                )  # Intent Name
                sheet.merge_cells(
                    start_row=first_row_intent,
                    end_row=row_num - 1,
                    start_column=2,
                    end_column=2,
                )  # Intent Display Name
                sheet.merge_cells(
                    start_row=first_row_intent,
                    end_row=row_num - 1,
                    start_column=3,
                    end_column=3,
                )  # Intent Description

        workbook.save(output_file)
        logger.info(f"Excel output saved to: {output_file}")

    def apply_recommendations_from_excel(
        self, input_file: str, agent_name: str, language_code: str
    ) -> None:
        """Applies intent recommendations from a formatted Excel file."""
        logger.info(f"Applying recommendations from file {input_file}")
        workbook = openpyxl.load_workbook(input_file)
        sheet = workbook.active

        intents_client = self.get_intents_client(agent_name=agent_name)
        entity_types = self.list_entity_types(agent_name, language_code)
        updated_intents: dict[str, list[dict[str, any]]] = {}
        updated_descriptions: dict[str, str] = {}

        current_intent_name: str | None = None
        current_intent_description: str | None = None
        training_phrase_current: str = ""
        training_phrase_new: str = ""

        for row_index, row in enumerate(
            sheet.iter_rows(min_row=2, values_only=True)
        ):  # Start from row 2
            (
                intent_name,
                intent_display_name,
                intent_description,
                training_phrase_id,
                repeat_count,
                current_tp_text,
                new_tp_text,
                recommendation,
                explanation,
                apply_change,
            ) = row

            if intent_name:
                if (
                    current_intent_name
                    and updated_intents.get(current_intent_name) is not None
                ):
                    self._process_intent_updates_excel(
                        updated_intents,
                        updated_descriptions,
                        intents_client,
                        agent_name,
                        language_code,
                        current_intent_name,
                        current_intent_description,
                        entity_types,
                    )

                current_intent_name = intent_name
                current_intent_description = intent_description
                updated_intents[current_intent_name] = updated_intents.get(
                    current_intent_name, []
                )
                updated_descriptions[current_intent_name] = intent_description
                training_phrase_current = ""
                training_phrase_new = ""

            if current_tp_text or new_tp_text or recommendation:
                training_phrase_current = current_tp_text if current_tp_text else ""
                training_phrase_new = new_tp_text if new_tp_text else ""

                if recommendation:
                    if apply_change == "X":
                        tp_data = {
                            "text": training_phrase_new,
                            "repeatCount": int(repeat_count) if repeat_count else 1,
                            "id": training_phrase_id
                            if recommendation != "ADD"
                            else None,
                            "original_text": training_phrase_current,
                            "recommendation": recommendation,
                        }
                        updated_intents[current_intent_name].append(tp_data)

        if current_intent_name and updated_intents.get(current_intent_name) is not None:
            self._process_intent_updates_excel(
                updated_intents,
                updated_descriptions,
                intents_client,
                agent_name,
                language_code,
                current_intent_name,
                current_intent_description,
                entity_types,
            )

    def _process_intent_updates_excel(
        self,
        updated_intents: dict[str, list[dict[str, any]]],
        updated_descriptions: dict[str, str],
        intents_client: dialogflowcx_v3.IntentsClient,
        agent_name: str,
        language_code: str,
        current_intent_name: str,
        current_intent_description: str,
        entity_types: list[dialogflowcx_v3.EntityType],
    ) -> None:
        """Helper function to process and apply intent updates from excel."""
        get_intent_request = dialogflowcx_v3.GetIntentRequest(
            name=current_intent_name, language_code=language_code
        )
        intent = intents_client.get_intent(get_intent_request)

        if updated_intents[current_intent_name]:
            updated_training_phrases_final = []
            original_tps = self.get_training_phrase_strings(intent)
            original_tp_objects = {tp.id: tp for tp in intent.training_phrases}
            tp_ids_to_remove = set()

            for updated_tp_data in updated_intents[current_intent_name]:
                recommendation_type = updated_tp_data["recommendation"]
                tp_id = updated_tp_data.get("id")
                # If no new training phrase is available, fall back to original training phrase
                new_text = updated_tp_data.get("text", updated_tp_data.get("original_text", ""))
                repeat_count = updated_tp_data.get("repeatCount", 1)

                if recommendation_type == "ADD":
                    if new_text:
                        updated_training_phrases_final.append(
                            {"text": new_text, "repeatCount": repeat_count, "id": None}
                        )                
                elif recommendation_type == "UPDATE" and tp_id:
                    updated_training_phrases_final.append(
                        {"text": new_text, "repeatCount": repeat_count, "id": tp_id}
                    )
                elif recommendation_type == "REMOVE" and tp_id:
                    tp_ids_to_remove.add(tp_id)
                elif recommendation_type == "RETAIN":
                    pass

                if recommendation_type != "":
                    logger.debug(
                        f"Applying recommendation type: {recommendation_type} for intent: {current_intent_name} and tp_id: {tp_id}"
                    )

            final_training_phrases_strings = []
            for tp_obj in intent.training_phrases:
                if tp_obj.id not in tp_ids_to_remove:
                    final_training_phrases_strings.append(
                        {
                            "text": self.format_training_phrase_string(tp_obj),
                            "repeatCount": tp_obj.repeat_count,
                            "id": tp_obj.id,
                        }
                    )

            final_training_phrases_strings.extend(updated_training_phrases_final)

            self.update_intent_training_phrases(
                intent,
                final_training_phrases_strings,
                agent_name,
                language_code,
                entity_types,
            )

        if intent.description != updated_descriptions[current_intent_name]:
            self.update_intent_description(
                intent,
                updated_descriptions[current_intent_name],
                agent_name,
                language_code,
            )


def main(
    agent_name: str,
    language_code: str,
    output_file: str | None = None,
    input_file: str | None = None,
    debug: bool | None = False,
    model_name: str | None = DEFAULT_LLM,
    vertex_ai_project: str | None = None,
    vertex_ai_location: str | None = None,
    gemini_timeout: int = 60,
) -> None:
    """Main function to orchestrate intent improvement process."""
    if debug:
        logging.getLogger().setLevel(logging.DEBUG)

    if not output_file and not input_file:
        logger.error("Either --output_file or --input_file must be specified.")
        return
    if output_file and input_file:
        logger.error("Only one of --output_file or --input_file can be specified.")
        return

    cabp = ConversationalAgentsBestPractice()

    intents = []
    all_recommendations = {}
    excel_written: bool = False
    error_occurred: bool = False

    try:
        logger.info("Loading intents...")
        intents = cabp.get_intents(agent_name, language_code)
        if intents is None:
            logger.error("Failed to retrieve intents. Exiting.")
            error_occurred = True
            return

        logger.info("Loading entity types and entities...")
        entity_types = cabp.list_entity_types(agent_name, language_code)
        logger.debug(f"Entity Types:\n{entity_types}")
        logger.info("Entity types and entities loaded.")

        if output_file:
            for intent in tqdm(
                intents,
                desc="Processing Intents",
                dynamic_ncols=True,
                position=0,
                leave=True,
            ):        
                # TODO: Remove
                if intent.name != "projects/ucds-testsystem/locations/europe-west3/agents/2febc044-4e2c-4e37-9195-24bf37d1f6d6/intents/9a169332-d65f-4b32-8e58-5e3d0b577c29":
                    continue                
                logger.info(f"Processing intent: {intent.display_name}")

                if not intent.description:
                    intent.description = cabp.generate_intent_description(
                        intent,
                        language_code,
                        model_name,
                        vertex_ai_project,
                        vertex_ai_location,
                        gemini_timeout,
                    )

                if intent.display_name == "Default Negative Intent":
                    logger.info(
                        f"Skipping recommendations for intent: {intent.display_name}"
                    )
                    continue

                recommendations = cabp.generate_training_phrase_recommendations(
                    intent,
                    language_code,
                    model_name,
                    entity_types,
                    vertex_ai_project,
                    vertex_ai_location,
                    gemini_timeout,
                )
                # Store original training phrase IDs with the recommendations
                original_tps = intent.training_phrases
                for rec in recommendations:
                    if "originalTrainingPhrase" in rec:
                        original_phrase_str = rec["originalTrainingPhrase"]
                        for tp in original_tps:
                            if cabp.format_training_phrase_string(tp) == original_phrase_str:
                                rec["originalTrainingPhraseId"] = tp.id
                                break

                all_recommendations[intent.name] = recommendations

                # TODO: remove
                #if intents.index(intent) == 5:
                #    break

            cabp.create_excel_output(
                intents, all_recommendations, output_file, agent_name, entity_types
            )
            excel_written = True            

        elif input_file:
            cabp.apply_recommendations_from_excel(input_file, agent_name, language_code)

    except KeyboardInterrupt:
        logger.warning("Command terminated by user (Ctrl+C).")
        error_occurred = True
    except Exception as e:
        logger.error(f"An unexpected error occurred: {e}")
        error_occurred = True
        traceback.print_tb(e.__traceback__)
    finally:
        if output_file and not excel_written and intents:
            logger.info("Writing Excel output due to program termination...")
            cabp.create_excel_output(
                intents, all_recommendations, output_file, agent_name, entity_types
            )
            logger.info(f"Excel output saved to: {output_file}")
        if error_occurred:
            sys.exit(1)
        else:
            sys.exit(0)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Improves Dialogflow CX intents.")
    parser.add_argument("agent_name", help="Dialogflow CX agent name")
    parser.add_argument("language_code", help="Language code for the agent")
    parser.add_argument(
        "--output_file", help="Path to output Excel file (output mode)"
    )
    parser.add_argument(
        "--input_file", help="Path to input Excel file (apply mode)"
    )
    parser.add_argument("--debug", action="store_true", help="Enable debug logging")
    parser.add_argument(
        "--model_name",
        default=DEFAULT_LLM,
        help=f"Gemini model name (default: {DEFAULT_LLM})",
    )
    parser.add_argument(
        "--vertex_ai_project",
        help="Vertex AI Project ID. Defaults to Dialogflow Agent Project ID if not specified.",
    )
    parser.add_argument(
        "--vertex_ai_location",
        help="Vertex AI Location. Defaults to Dialogflow Agent Location if not specified.",
    )
    parser.add_argument(
        "--gemini_timeout",
        type=int,
        default=LLM_TIMEOUT_SECONDS,
        help=f"Timeout in seconds for Gemini API calls (default: {LLM_TIMEOUT_SECONDS} seconds)",
    )

    args = parser.parse_args()

    if not args.vertex_ai_project:
        args.vertex_ai_project = args.agent_name.split("/")[1]

    if not args.vertex_ai_location:
        args.vertex_ai_location = args.agent_name.split("/")[3]

    main(
        args.agent_name,
        args.language_code,
        args.output_file,
        args.input_file,
        args.debug,
        args.model_name,
        args.vertex_ai_project,
        args.vertex_ai_location,
        args.gemini_timeout,
    )